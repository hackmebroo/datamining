import heapq
import matplotlib.pyplot as plt
import numpy
import xlrd
import numpy as np
import xlsxwriter as xlsxwriter
from openpyxl import load_workbook


def create_dataset_labels(path):                # 获得数据集和标签集
    table = xlrd.open_workbook(path).sheets()[0]  # 获取第一个sheet表
    row = table.nrows  # 行数
    col = table.ncols  # 列数
    datamatrix = np.zeros((row, col))  # 生成一个nrows行ncols列，且元素均为0的初始矩阵
    for x in range(col):
        cols = np.matrix(table.col_values(x))  # 把list转换为矩阵进行矩阵操作
        datamatrix[:, x] = cols  # 按列把数据存进矩阵中

    wb = load_workbook(filename=r'E:\\dataset\\kNN.xlsx')
    # 读取名字为Sheet2的表
    ws = wb["Sheet2"]
    labels = []
    for row_A in range(1, 1000):
        # 遍历第1列
        a1 = ws.cell(row=row_A, column=1).value
        if a1:
            # 写入数组1
            labels.append(a1)
    return datamatrix, labels


def get_k_means(path):
    table = xlrd.open_workbook(path).sheets()[0]  # 获取第一个sheet表
    row = table.nrows  # 行数
    col = table.ncols  # 列数
    k_means_set = np.zeros((row, col))  # 生成一个nrows行ncols列，且元素均为0的初始矩阵
    for x in range(col):
        cols = np.matrix(table.col_values(x))  # 把list转换为矩阵进行矩阵操作
        k_means_set[:, x] = cols  # 按列把数据存进矩阵中

    table = xlrd.open_workbook(path).sheets()[1]  # 获取第二个sheet表
    row = table.nrows  # 行数
    col = table.ncols  # 列数
    k_means_set_index = np.zeros((row, col))  # 生成一个nrows行ncols列，且元素均为0的初始矩阵
    for x in range(col):
        cols = np.matrix(table.col_values(x))  # 把list转换为矩阵进行矩阵操作
        k_means_set_index[:, x] = cols  # 按列把数据存进矩阵中
    return k_means_set, k_means_set_index


def create_W(labels, k_means_set_index):
    W = np.zeros((500, 500))
    for i in range(500):
        for j in range(500):
            for x in range(4):
                if j == k_means_set_index[i][x] and labels[i] == labels[j]:
                    W[i][j] = 1
                    break
                elif i == k_means_set_index[j][x] and labels[i] == labels[j]:
                    W[i][j] = 1
                    break
                elif j == k_means_set_index[i][x] and labels[i] != labels[j]:
                    W[i][j] = -1
                    break
                elif i == k_means_set_index[j][x] and labels[i] != labels[j]:
                    W[i][j] = -1
                    break
                else:
                    W[i][j] = 0
    return W


def create_L(W):
    D = np.zeros((500, 500))
    for i in range(500):
        D[i][i] = sum(W[:][i])
    L = D - W
    return L


def cal_XLX(X, L):
    X1 = X.transpose()
    A1 = numpy.dot(X1, L)
    A2 = numpy.dot(A1, X)
    return A2


def cal_eig(A):
    eigenvalue, eigenvector = np.linalg.eig(A)
    sorted_indices = np.argsort(eigenvalue)
    top1_evecs = eigenvector[:, sorted_indices[:-1 - 1:-1]]
    top2_evecs = eigenvector[:, sorted_indices[:-2 - 1:-1]]
    return top1_evecs, top2_evecs


def cal_Y(X, P_1, P_2):
    X1 = X.transpose()
    P1 = P_1.transpose()
    P2 = P_2.transpose()
    Y_1 = numpy.dot(P1, X1)
    Y_2 = numpy.dot(P2, X1)
    Y1 = Y_1.transpose()
    Y2 = Y_2.transpose()
    return Y1, Y2


datafile = u'E:\\dataset\\dataset.xlsx'
datafile1 = u'E:\\dataset\\k_means_set.xlsx'
X, labels = create_dataset_labels(datafile)
k_means_set, k_means_set_index = get_k_means(datafile1)
W = create_W(labels, k_means_set_index)
L = create_L(W)
A = cal_XLX(X, L)
P1, P2 = cal_eig(A)
Y1, Y2 = cal_Y(X, P1, P2)
# tmp = np.zeros((500, 1))
for i in range(500):
    if i < 100:
        plt.scatter(Y1[i, 0], 0, marker='o', color='red')
    elif i < 300:
        plt.scatter(Y1[i, 0], 0, marker='o', color='yellow')
    elif i < 500:
        plt.scatter(Y1[i, 0], 0, marker='o', color='blue')

# plt.scatter(Y2[:, 0], Y2[:, 1], marker='o')
# plt.savefig("DNE二维.png")
# plt.show()

# plt.scatter(Y1[:, 0], tmp, marker='o')
plt.savefig("DNE一维.png")
plt.show()



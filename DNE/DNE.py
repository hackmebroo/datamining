import heapq

import numpy
import xlrd
import numpy as np
import xlsxwriter as xlsxwriter
from openpyxl import load_workbook


def create_dataset_labels(path):                # 获得数据集和标签集
    table = xlrd.open_workbook(path).sheets()[0]  # 获取第一个sheet表
    row = table.nrows  # 行数
    col = table.ncols  # 列数
    datamatrix = np.zeros((row, col))  # 生成一个nrows行ncols列，且元素均为0的初始矩阵
    for x in range(col):
        cols = np.matrix(table.col_values(x))  # 把list转换为矩阵进行矩阵操作
        datamatrix[:, x] = cols  # 按列把数据存进矩阵中

    wb = load_workbook(filename=r'E:\\dataset\\kNN.xlsx')
    # 读取名字为Sheet2的表
    ws = wb["Sheet2"]
    labels = []
    for row_A in range(1, 1000):
        # 遍历第1列
        a1 = ws.cell(row=row_A, column=1).value
        if a1:
            # 写入数组1
            labels.append(a1)
    return datamatrix, labels


def measure_dist(a, b):                 # 求两点之间距离
    dist = numpy.linalg.norm(a - b)
    return dist


def k_set(dataset, k):                  # 求每个点的k邻近集
    dataset_size = dataset.shape[0]
    k_means_set = []
    k_means_set_index = []
    for i in range(dataset_size):
        dist_tmp = []
        for j in range(dataset_size):
            dist = measure_dist(dataset[i], dataset[j])
            dist_tmp.append(dist)
        min_k_set = heapq.nsmallest(k, dist_tmp)                # 求距离最小的k个值
        min_k_set_index = list(map(dist_tmp.index, heapq.nsmallest(k, dist_tmp)))            # 求这k个值的索引
        k_means_set.append(min_k_set)
        k_means_set_index.append(min_k_set_index)
    workbook = xlsxwriter.Workbook('k_means_set.xlsx')  # 生成表格
    worksheet1 = workbook.add_worksheet(u'sheet1')
    worksheet2 = workbook.add_worksheet(u'sheet2')
    for i in range(len(k_means_set)):
        for j in range(k):
            worksheet1.write(i, j, k_means_set[i][j])
    for i in range(len(k_means_set_index)):
        for j in range(k):
            worksheet2.write(i, j, k_means_set_index[i][j])
    workbook.close()
    return k_means_set


datafile = u'E:\\dataset\\dataset.xlsx'
X, labels = create_dataset_labels(datafile)
k_means_set = k_set(X, 4)
print(k_means_set)


import random
import numpy
import numpy as np
import operator
import xlrd
from numpy import tile
from openpyxl import load_workbook
import matplotlib.pyplot as plt


def select(num, dataSet):
    select_set = []
    select_set_index = []
    init_num = dataSet.__len__()
    ran = random.randint(0, init_num)
    # dataSet[ran].select()
    select_set.append(dataSet[ran])          # 选取最开始的样本点
    select_set_index.append(ran)

    round = 1
    while round < num:
        tmp_max_distance = []
        for i in range(init_num):
            # if dataSet[i].label == 1:
            # if i < 300:
                # continue
            # else:
                tmp = []
                for selected in select_set:
                    # dist = measure_distance(dataSet[i], dataSet[selected])
                    dist = numpy.sqrt(numpy.sum(numpy.square(dataSet[i] - selected)))
                    tmp.append(dist)
                tmp_max_distance.append((min(tmp), i))      # (距离，索引)
        # print(tmp_max_distance)

        tmp_max = 0
        tmp_tuple = (0, -1)
        for tmp_dis_tuple in tmp_max_distance:
            if tmp_dis_tuple[0] > tmp_max:
                tmp_tuple = tmp_dis_tuple
                tmp_max = tmp_dis_tuple[0]
        if tmp_tuple[1] != -1:
            # dataSet[tmp_tuple[1]].select()
            select_set.append(dataSet[tmp_tuple[1]])
            select_set_index.append(tmp_tuple[1])
        # print(tmp_tuple)
        round += 1
    return select_set, select_set_index


def createDataSet(path):
    """
    函数作用：构建一组训练数据（训练样本），共4个样本
    同时给出了这4个样本的标签，及labels
    """
    table1 = xlrd.open_workbook(path).sheets()[0]  # 获取第一个sheet表
    row1 = table1.nrows  # 行数
    col1 = table1.ncols  # 列数
    # 样本group:
    group = np.zeros((row1, col1))  # 生成一个nrows行ncols列，且元素均为0的初始矩阵
    for x in range(col1):
        cols1 = np.matrix(table1.col_values(x))  # 把list转换为矩阵进行矩阵操作
        group[:, x] = cols1  # 按列把数据存进矩阵中
    # 标签labels:
    # 读取路径
    wb = load_workbook(filename=r'E:\\dataset\\kNN.xlsx')
    # 读取名字为Sheet2的表
    ws = wb["Sheet2"]
    labels = []
    for row_A in range(1, 10000):
        # 遍历第1行到10000行，第1列
        a1 = ws.cell(row=row_A, column=1).value
        if a1:
            # 写入数组1
            labels.append(a1)
    # print(group)
    # print(labels)
    return group, labels


def classify0(inX, dataset, labels, k):
    """
    inX 是输入的测试样本，是一个[x, y]样式的
    dataset 是训练样本集
    labels 是训练样本标签
    k 是top k最相近的
    """
    # shape返回矩阵的[行数，列数]，
    # 那么shape[0]获取数据集的行数，
    # 行数就是样本的数量
    dataSetSize = dataset.shape[0]

    """
    下面的求距离过程就是按照欧氏距离的公式计算的。
    即 根号(x^2+y^2)
    """
    # tile属于numpy模块下边的函数
    # tile（A, reps）返回一个shape=reps的矩阵，矩阵的每个元素是A
    # 比如 A=[0,1,2] 那么，tile(A, 2)= [0, 1, 2, 0, 1, 2]
    # tile(A,(2,2)) = [[0, 1, 2, 0, 1, 2],
    #                  [0, 1, 2, 0, 1, 2]]
    # tile(A,(2,1,2)) = [[[0, 1, 2, 0, 1, 2]],
    #                    [[0, 1, 2, 0, 1, 2]]]
    # 上边那个结果的分开理解就是：
    # 最外层是2个元素，即最外边的[]中包含2个元素，类似于[C,D],而此处的C=D，因为是复制出来的
    # 然后C包含1个元素，即C=[E],同理D=[E]
    # 最后E包含2个元素，即E=[F,G],此处F=G，因为是复制出来的
    # F就是A了，基础元素
    # 综合起来就是(2,1,2)= [C, C] = [[E], [E]] = [[[F, F]], [[F, F]]] = [[[A, A]], [[A, A]]]
    # 这个地方就是为了把输入的测试样本扩展为和dataset的shape一样，然后就可以直接做矩阵减法了。
    # 比如，dataset有4个样本，就是4*2的矩阵，输入测试样本肯定是一个了，就是1*2，为了计算输入样本与训练样本的距离
    # 那么，需要对这个数据进行作差。这是一次比较，因为训练样本有n个，那么就要进行n次比较；
    # 为了方便计算，把输入样本复制n次，然后直接与训练样本作矩阵差运算，就可以一次性比较了n个样本。
    # 比如inX = [0,1],dataset就用函数返回的结果，那么
    # tile(inX, (4,1))= [[ 0.0, 1.0],
    #                    [ 0.0, 1.0],
    #                    [ 0.0, 1.0],
    #                    [ 0.0, 1.0]]
    # 作差之后
    # diffMat = [[-1.0,-0.1],
    #            [-1.0, 0.0],
    #            [ 0.0, 1.0],
    #            [ 0.0, 0.9]]
    diffMat = tile(inX, (dataSetSize, 1)) - dataset

    # diffMat就是输入样本与每个训练样本的差值，然后对其每个x和y的差值进行平方运算。
    # diffMat是一个矩阵，矩阵**2表示对矩阵中的每个元素进行**2操作，即平方。
    # sqDiffMat = [[1.0, 0.01],
    #              [1.0, 0.0 ],
    #              [0.0, 1.0 ],
    #              [0.0, 0.81]]
    sqDiffMat = diffMat ** 2

    # axis=1表示按照横轴，sum表示累加，即按照行进行累加。
    # sqDistance = [[1.01],
    #               [1.0 ],
    #               [1.0 ],
    #               [0.81]]
    sqDistance = sqDiffMat.sum(axis=1)

    # 对平方和进行开根号
    distance = sqDistance ** 0.5

    # 按照升序进行快速排序，返回的是原数组的下标。
    # 比如，x = [30, 10, 20, 40]
    # 升序排序后应该是[10,20,30,40],他们的原下标是[1,2,0,3]
    # 那么，numpy.argsort(x) = [1, 2, 0, 3]
    sortedDistIndicies = distance.argsort()

    # 存放最终的分类结果及相应的结果投票数
    classCount = {}

    # 投票过程，就是统计前k个最近的样本所属类别包含的样本个数
    for i in range(k):
        # index = sortedDistIndicies[i]是第i个最相近的样本下标
        # voteIlabel = labels[index]是样本index对应的分类结果('A' or 'B')
        voteIlabel = labels[sortedDistIndicies[i]]
        # classCount.get(voteIlabel, 0)返回voteIlabel的值，如果不存在，则返回0
        # 然后将票数增1
        classCount[voteIlabel] = classCount.get(voteIlabel, 0) + 1

    # 把分类结果进行排序，然后返回得票数最多的分类结果
    sortedClassCount = sorted(classCount.items(), key=operator.itemgetter(1), reverse=True)
    return sortedClassCount[0][0]


if __name__ == "__main__":
    # 导入数据
    datafile = u'E:\\dataset\\kNN.xlsx'
    dataset, labels = createDataSet(datafile)
    plt.scatter(dataset[0:299, 0], dataset[0:299, 1], color='red')
    plt.scatter(dataset[300:999, 0], dataset[300:999, 1], color='yellow')
    inX, inX_index = select(100, dataset)
    # 简单分类
    tmp = 0
    tmp1 = np.zeros((1, 1))
    i = 0
    while i < 100:
        className = classify0(inX[i], dataset, labels, 5)
        if className == 'A':
            tmp1 = dataset[inX_index[i]]
            plt.scatter(tmp1[0], tmp1[1], marker='+', color='black')
            if inX_index[i] >= 300:
                print('sample[%d]分类错误' % i)
                tmp += 1
        else:
            tmp1 = dataset[inX_index[i]]
            plt.scatter(tmp1[0], tmp1[1], marker='*', color='blue')
            if inX_index[i] < 300:
                print('sample[%d]分类错误' % i)
                tmp += 1
        print('the class of test sample[%d] is %s' % (i, className))
        i += 1
    print('分类正确率为：%f' % ((100 - tmp)/100))

    # plt.scatter(dataset[0:299, 0], dataset[0:299, 1], color='red')
    # plt.scatter(dataset[300:999, 0], dataset[300:999, 1], color='yellow')
    plt.savefig("kNN.png")
    plt.show()
